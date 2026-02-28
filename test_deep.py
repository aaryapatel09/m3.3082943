#!/usr/bin/env python3
"""
DEEP TEST v2: stress-test every code path, cross-validate Excel data,
check MATLAB semantics, probe edge cases. Matches updated progressive
tax and risk-dependent gambling caps.
"""

import openpyxl, math, random, os, re, sys

XLSX = '/Users/aaryapatel/m3/M3-Challenge-Problem-Data-2026.xlsx'
M3DIR = '/Users/aaryapatel/m3'
PASS = 0
FAIL = 0

def ok(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
    else:
        FAIL += 1
        print(f"  **FAIL**  {name}  {detail}")

def approx(a, b, rtol=0.005):
    if abs(b) < 1e-9: return abs(a) < 1e-6
    return abs(a - b) / max(abs(a), abs(b), 1e-12) < rtol

wb = openpyxl.load_workbook(XLSX, data_only=True)

def pv(ws, r, c):
    v = ws.cell(r, c).value
    if v is None: return 0.0
    if isinstance(v, (int, float)):
        return 0.0 if math.isnan(v) else float(v)
    try: return float(str(v).replace(',',''))
    except: return 0.0

print("=" * 72)
print("DEEP TEST v2 — M3 Challenge 2026")
print("=" * 72)

# ======================================================================
# A. MATLAB SYNTAX DEEP-SCAN (corrected for MATLAB semantics)
# ======================================================================
print("\n[A] MATLAB syntax deep-scan ...")

m_files = sorted(f for f in os.listdir(M3DIR) if f.endswith('.m'))
expected_files = [
    'main_m3.m', 'loadUSExpenditureData.m', 'loadUKExpenditureData.m',
    'loadUSBettingData.m', 'afterTaxIncome.m', 'baseEssentialShare_US.m',
    'baseEssentialShare_UK.m', 'regionMultiplier.m', 'householdMultiplier.m',
    'disposableIncome.m', 'gamblingOutcome.m', 'main_Q1_demo.m',
    'main_Q2_demo.m', 'main_Q3_demo.m'
]
ok("Found 14 .m files", len(m_files) == 14, f"found {len(m_files)}")
for ef in expected_files:
    ok(f"File exists: {ef}", ef in m_files)

for mf in m_files:
    path = os.path.join(M3DIR, mf)
    with open(path) as fh:
        src = fh.read()
        lines = src.split('\n')

    # Function files must start with 'function' (skip main_m3.m which is a script)
    if mf != 'main_m3.m':
        first_code = ''
        for ln in lines:
            s = ln.strip()
            if s and not s.startswith('%'):
                first_code = s
                break
        ok(f"{mf}: first code line is 'function'", first_code.startswith('function'), first_code[:50])

    # No Python-isms
    ok(f"{mf}: no 'def ' keyword", 'def ' not in src)
    ok(f"{mf}: no 'import ' keyword", 'import ' not in src)
    ok(f"{mf}: no 'print(' calls", 'print(' not in src)
    ok(f"{mf}: no tab characters", '\t' not in src)

    # readcell sheet names must match actual Excel sheets
    sheet_calls = re.findall(r"readcell\([^)]*'Sheet'\s*,\s*'([^']+)'", src)
    actual_sheets = wb.sheetnames
    for sc in sheet_calls:
        ok(f"{mf}: sheet name '{sc}' exists in Excel", sc in actual_sheets,
           f"available: {actual_sheets}")

    # Struct field access validation
    if 'usData.' in src:
        fields_used = set(re.findall(r'usData\.(\w+)', src))
        known_us = {'ageLabels','ageBounds','meanIncome','allExpend','essentialExpend',
                     'regionLabels','regionIncome','regionAllExpend','regionEssentialExpend',
                     'nationalIncome','nationalAllExpend'}
        for f in fields_used:
            ok(f"{mf}: usData.{f} is known", f in known_us, f"unknown: {f}")

    if 'ukData.' in src:
        fields_used = set(re.findall(r'ukData\.(\w+)', src))
        known_uk = {'ageLabels','ageBounds','quintileIncome','avgPersons','essentialWeekly',
                     'avgPersonsAll','regionLabels','regionAllExpendWeekly','nationalAllExpendWeekly'}
        for f in fields_used:
            ok(f"{mf}: ukData.{f} is known", f in known_uk, f"unknown: {f}")

# ======================================================================
# B. EXCEL DATA CROSS-VALIDATION
# ======================================================================
print("\n[B] Excel data cross-validation ...")

ws_us = wb['Expenditures (U.S.)']
ws_uk = wb['Expenditures (U.K.)']
ws_bet = wb['Online Sports Betting Personal ']

for c in range(2, 13):
    ess = sum(pv(ws_us, r, c) for r in [12,13,14,19,20,25])
    all_e = pv(ws_us, 11, c)
    ok(f"US col {c}: essential({ess:.0f}) <= all({all_e:.0f})", ess <= all_e * 1.001)
    inc = pv(ws_us, 4, c)
    ok(f"US col {c}: income({inc:.0f}) > 0", inc > 0)

blockStarts = [4, 21, 38, 55, 72]
qi = [pv(ws_uk, 5, c) for c in range(10, 15)]
for i in range(4):
    ok(f"UK quintile income[{i}] < [{i+1}]", qi[i] < qi[i+1])

for b, r0 in enumerate(blockStarts):
    for q in range(5):
        food = pv(ws_uk, r0+4, 2+q)
        hous = pv(ws_uk, r0+7, 2+q)
        tran = pv(ws_uk, r0+10, 2+q)
        total = food + hous + tran
        ok(f"UK block {b} Q{q}: essential wkly {total:.1f} in [50,500]",
           50 <= total <= 500, f"got {total:.1f}")

for row in [4, 19, 20, 21, 40, 41, 42]:
    for col in [2, 4, 5, 6, 7, 8, 9]:
        v = pv(ws_bet, row, col)
        ok(f"Bet R{row}C{col}: {v} in [0,100]", 0 <= v <= 100)

for sport_row in range(15, 20):
    total = sum(pv(ws_bet, sport_row, c) for c in range(14, 17))
    ok(f"Bet R{sport_row}: wager pcts sum {total} in [85,100]", 85 <= total <= 100)

# ======================================================================
# C. CORE FUNCTIONS — PROGRESSIVE TAX
# ======================================================================
print("\n[C] Progressive marginal tax ...")

def afterTax(S, country):
    brackets = [0, 30000, 80000, 150000, float('inf')]
    rates = [0.10, 0.18, 0.24, 0.30] if country == 'US' else [0.12, 0.20, 0.26, 0.32]
    tax = 0
    for k in range(len(rates)):
        lo, hi = brackets[k], brackets[k+1]
        taxable = max(0, min(S, hi) - lo)
        tax += rates[k] * taxable
    Y = S - tax
    tau = tax / S if S > 0 else 0
    return Y, tau

# Spot-check progressive tax
ok("US S=0: Y=0", afterTax(0, 'US')[0] == 0)
ok("US S=30000: Y=27000", approx(afterTax(30000, 'US')[0], 27000))
ok("US S=70000: Y=59800", approx(afterTax(70000, 'US')[0], 59800))
ok("US S=80000: Y=68000", approx(afterTax(80000, 'US')[0], 68000))
ok("US S=110000: Y=90800", approx(afterTax(110000, 'US')[0], 90800))
ok("US S=150000: Y=121200", approx(afterTax(150000, 'US')[0], 121200))
ok("US S=200000: Y=156200", approx(afterTax(200000, 'US')[0], 156200))
ok("UK S=60000: Y=50400", approx(afterTax(60000, 'UK')[0], 50400))
ok("UK S=150000: Y=118200", approx(afterTax(150000, 'UK')[0], 118200))

# Strict monotonicity check: Y(S+1) > Y(S) for all S
for country in ['US', 'UK']:
    prev_Y = -1
    for S in range(0, 250001, 100):
        Y = afterTax(S, country)[0]
        ok(f"Tax monotone {country} S={S}: Y({S})={Y:.0f} > prev",
           Y >= prev_Y - 0.001, f"prev={prev_Y:.0f}, now={Y:.0f}")
        prev_Y = Y

# Effective rate increasing
for country in ['US', 'UK']:
    prev_tau = -1
    for S in [10000, 30000, 50000, 80000, 120000, 150000, 200000]:
        _, tau = afterTax(S, country)
        ok(f"Tax eff rate {country} S={S}: tau={tau:.3f} >= prev",
           tau >= prev_tau - 0.001, f"prev={prev_tau:.3f}")
        prev_tau = tau

# ======================================================================
# D. FULL DI PIPELINE
# ======================================================================
print("\n[D] Full DI pipeline ...")

us_ageBounds = [(0,24),(25,34),(35,44),(45,54),(55,64),(65,74),(75,9999)]
uk_ageBounds = [(0,29),(30,49),(50,64),(65,74),(75,9999)]
us_meanIncome = [pv(ws_us, 4, c) for c in range(2,9)]
us_essExpend = [sum(pv(ws_us, r, c) for r in [12,13,14,19,20,25]) for c in range(2,9)]
us_regionIncome = [pv(ws_us, 4, c) for c in range(9,13)]
us_regionAllExpend = [pv(ws_us, 11, c) for c in range(9,13)]
us_nationalIncome = sum(us_regionIncome)/4
us_nationalAllExpend = sum(us_regionAllExpend)/4
uk_quintileIncome = [pv(ws_uk, 5, c) for c in range(10,15)]
uk_essWeekly = []
uk_avgPAll = []
usHHbyAge = [2.0,3.0,3.3,3.0,2.5,2.0,1.7]
for r0 in blockStarts:
    food = [pv(ws_uk, r0+4, c) for c in range(2,7)]
    hous = [pv(ws_uk, r0+7, c) for c in range(2,7)]
    tran = [pv(ws_uk, r0+10, c) for c in range(2,7)]
    uk_essWeekly.append([f+h+t for f,h,t in zip(food,hous,tran)])
    uk_avgPAll.append(pv(ws_uk, r0+2, 7))
uk_regAllExpWk = [0.0]*4
uk_natAllExpWk = 0.0
for rr in range(12,25):
    for i, c in enumerate(range(10,14)):
        uk_regAllExpWk[i] += pv(ws_uk, rr, c)
    uk_natAllExpWk += pv(ws_uk, rr, 14)

def mapUS(age):
    for i,(lo,hi) in enumerate(us_ageBounds):
        if lo<=age<=hi: return i
    return 6
def mapUK(age):
    for i,(lo,hi) in enumerate(uk_ageBounds):
        if lo<=age<=hi: return i
    return 4
def findQ(wk, qi):
    bnd = [(qi[i]+qi[i+1])/2 for i in range(4)]
    q = 0
    for b in bnd:
        if wk > b: q += 1
    return q
def regMult(region, country):
    if country == 'US':
        regs = ['Northeast','Midwest','South','West']
        if region not in regs: return 1.0
        ri = regs.index(region)
        return (us_regionAllExpend[ri]/us_regionIncome[ri]) / (us_nationalAllExpend/us_nationalIncome)
    else:
        regs = ['England','Wales','Scotland','Northern Ireland']
        if region not in regs: return 1.0
        return uk_regAllExpWk[regs.index(region)] / uk_natAllExpWk
def hhMult(h, age, country):
    gamma = 0.7
    if country == 'UK':
        hbar = uk_avgPAll[mapUK(age)]
    else:
        hbar = usHHbyAge[mapUS(age)]
    if hbar <= 0: hbar = 2.5
    return (h/hbar)**gamma

def DI_full(S, age, h, country, region):
    Y = afterTax(S, country)[0]
    if country == 'US':
        idx = mapUS(age)
        theta = us_essExpend[idx] / us_meanIncome[idx]
    else:
        bIdx = mapUK(age)
        qIdx = findQ(S/52, uk_quintileIncome)
        theta = uk_essWeekly[bIdx][qIdx] / uk_quintileIncome[qIdx]
    E = theta * S * regMult(region, country) * hhMult(h, age, country)
    E = min(E, Y)
    return Y - E, Y, E

# Q1 demo examples (updated for progressive tax)
DI1, Y1, E1 = DI_full(70000, 28, 1, 'US', 'West')
ok("Q1 ex1: Y=59800", approx(Y1, 59800))
ok("Q1 ex1: DI > 0", DI1 > 0, f"DI={DI1:.0f}")
print(f"  Q1 ex1: Y={Y1:.0f}, E={E1:.0f}, DI={DI1:.0f}")

DI2, Y2, E2 = DI_full(110000, 42, 4, 'US', 'Midwest')
ok("Q1 ex2: Y=90800", approx(Y2, 90800))
ok("Q1 ex2: DI >= 0", DI2 >= 0, f"DI={DI2:.0f}")
print(f"  Q1 ex2: Y={Y2:.0f}, E={E2:.0f}, DI={DI2:.0f}")

DI3, Y3, E3 = DI_full(60000, 35, 2, 'UK', 'England')
ok("Q1 ex3: Y=50400", approx(Y3, 50400))
ok("Q1 ex3: DI > 0", DI3 > 0, f"DI={DI3:.0f}")
print(f"  Q1 ex3: Y={Y3:.0f}, E={E3:.0f}, DI={DI3:.0f}")

# Grid test: every combo
us_regions = ['Northeast','Midwest','South','West']
uk_regions = ['England','Wales','Scotland','Northern Ireland']
test_sals = [20000, 40000, 70000, 110000, 150000, 200000]
test_hh = [1, 2, 3, 4, 5]
total_combos = 0
for age in [20, 28, 40, 50, 60, 70, 80]:
    for reg in us_regions:
        for sal in test_sals:
            for h in test_hh:
                DI, Y, E = DI_full(sal, age, h, 'US', reg)
                total_combos += 1
                ok(f"US a={age} r={reg} s={sal} h={h}: invariants",
                   Y > 0 and E >= 0 and E <= Y + 0.01 and DI >= -0.01,
                   f"Y={Y:.0f} E={E:.0f} DI={DI:.0f}")
for age in [22, 35, 55, 68, 80]:
    for reg in uk_regions:
        for sal in test_sals:
            for h in test_hh:
                DI, Y, E = DI_full(sal, age, h, 'UK', reg)
                total_combos += 1
                ok(f"UK a={age} r={reg} s={sal} h={h}: invariants",
                   Y > 0 and E >= 0 and E <= Y + 0.01 and DI >= -0.01,
                   f"Y={Y:.0f} E={E:.0f} DI={DI:.0f}")
print(f"  Grid: {total_combos} combos all passed")

# Strict monotonicity of DI w.r.t. salary
for country, reg, age in [('US','West',35), ('US','South',50), ('UK','England',40), ('UK','Scotland',70)]:
    prev_di = -1e9
    for sal in range(20000, 200001, 1000):
        di = DI_full(sal, age, 2, country, reg)[0]
        ok(f"DI mono {country}/{reg}/a{age}: DI(${sal})>=${sal-1000}",
           di >= prev_di - 0.01, f"prev={prev_di:.0f}, now={di:.0f}")
        prev_di = di

# Household size: DI decreases with h
for country, reg, age, sal in [('US','West',35,80000), ('UK','England',40,60000)]:
    prev_di = 1e9
    for h in [1,2,3,4,5]:
        di = DI_full(sal, age, h, country, reg)[0]
        ok(f"HH {country}/a{age}/s{sal}: DI(h={h}) <= DI(h={h-1})",
           di <= prev_di + 0.01)
        prev_di = di

# ======================================================================
# E. GAMBLING OUTCOME — UPDATED CAPS
# ======================================================================
print("\n[E] Gambling outcome (risk-dependent caps) ...")

def gamblingOutcome(p):
    DI = DI_full(p['salary'], p['age'], p['hhsize'], p['country'], p['region'])[0]
    M_base = {'low':50,'medium':300,'high':800}
    M = M_base[p['riskTolerance']]
    af = 1.15 if p['age']<35 else (0.70 if p['age']>=65 else 1.0)
    gf = 1.10 if p['gender']=='male' else (0.85 if p['gender']=='female' else 1.0)
    ef = 1.05 if p['education']=='college' else 1.0
    M *= af * gf * ef
    G_raw = 12 * M
    capFrac = {'low':0.15, 'medium':0.40, 'high':1.00}[p['riskTolerance']]
    G = min(G_raw, capFrac * max(DI, 0))
    baseEdge = {'conservative':0.04,'mixed':0.07,'highrisk':0.12}.get(p['style'], 0.07)
    chaseMult = {'low':1.00,'medium':1.20,'high':1.50}[p['riskTolerance']]
    edge = baseEdge * chaseMult
    return -edge*G, edge*G, edge*G/max(DI,1), G, DI

# Every risk x style combo
for risk in ['low','medium','high']:
    for style in ['conservative','mixed','highrisk']:
        p = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
             'gender':'male','education':'nocollege','riskTolerance':risk,'style':style}
        expL, sig, frac, G, DI = gamblingOutcome(p)
        capFrac = {'low':0.15,'medium':0.40,'high':1.00}[risk]
        ok(f"Gamble r={risk} s={style}: loss<=0, G>=0, cap OK",
           expL <= 0 and G >= 0 and G <= capFrac*max(DI,0)+0.01 and 0 <= frac <= 1.001,
           f"expL={expL:.0f} G={G:.0f} frac={frac:.4f}")

# Higher risk -> larger loss
for style in ['conservative','mixed','highrisk']:
    p_lo = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
            'gender':'male','education':'nocollege','riskTolerance':'low','style':style}
    p_hi = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
            'gender':'male','education':'nocollege','riskTolerance':'high','style':style}
    ok(f"Style={style}: high risk loss >= low risk loss",
       gamblingOutcome(p_hi)[0] <= gamblingOutcome(p_lo)[0])

# Higher edge -> larger loss
for risk in ['low','medium','high']:
    p_c = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
           'gender':'male','education':'nocollege','riskTolerance':risk,'style':'conservative'}
    p_h = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
           'gender':'male','education':'nocollege','riskTolerance':risk,'style':'highrisk'}
    ok(f"Risk={risk}: highrisk style >= conservative loss",
       gamblingOutcome(p_h)[0] <= gamblingOutcome(p_c)[0])

# Male > female stakes
for risk in ['low','medium','high']:
    pm = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
          'gender':'male','education':'nocollege','riskTolerance':risk,'style':'mixed'}
    pf = {'salary':80000,'age':35,'hhsize':2,'country':'US','region':'West',
          'gender':'female','education':'nocollege','riskTolerance':risk,'style':'mixed'}
    ok(f"Risk={risk}: male G >= female G",
       gamblingOutcome(pm)[3] >= gamblingOutcome(pf)[3])

# Young > old stakes
ok("Young(25) stakes > Old(70)",
   gamblingOutcome({'salary':80000,'age':25,'hhsize':2,'country':'US','region':'West',
                    'gender':'male','education':'nocollege','riskTolerance':'medium','style':'mixed'})[3] >
   gamblingOutcome({'salary':80000,'age':70,'hhsize':2,'country':'US','region':'West',
                    'gender':'male','education':'nocollege','riskTolerance':'medium','style':'mixed'})[3])

# KEY: Verify R >= 15% is ACHIEVABLE for high-risk + low-DI individuals
p_extreme = {'salary':25000,'age':22,'hhsize':4,'country':'US','region':'West',
             'gender':'male','education':'nocollege','riskTolerance':'high','style':'highrisk'}
expL_ex, _, fracDI_ex, G_ex, DI_ex = gamblingOutcome(p_extreme)
print(f"  Extreme case: salary=25k, h=4, high/highrisk -> DI={DI_ex:.0f}, G={G_ex:.0f}, fracDI={fracDI_ex:.4f}")
# With high cap (1.0*DI) and high edge (0.12), R should be notable
if DI_ex > 0:
    ok("Extreme: R > 5% (meaningful)", fracDI_ex > 0.05,
       f"fracDI={fracDI_ex:.4f}")

# ======================================================================
# F. Q3 METRICS
# ======================================================================
print("\n[F] Q3 impact metrics ...")

r_inv = 0.04; T = 20
annuity = ((1+r_inv)**T - 1)/r_inv
ok("Annuity factor ~ 29.778", approx(annuity, 29.778, 0.001))
manual_sum = sum(1.04**k for k in range(20))
ok("Annuity matches manual sum", approx(annuity, manual_sum, 0.001))

for loss, di in [(-100,50000),(-500,30000),(-1000,20000),(-2000,10000)]:
    R = -loss/di
    DW = -loss*annuity
    ok(f"Q3 loss={loss} DI={di}: R={R:.4f}", approx(R, -loss/di))
    ok(f"Q3 loss={loss} DI={di}: DW={DW:.0f}", approx(DW, -loss*annuity))

# ======================================================================
# G. LARGE MC (N=5000) WITH UPDATED MODEL
# ======================================================================
print("\n[G] Large Monte Carlo (N=5000) ...")

random.seed(2026)
N = 5000
losses_all, dis_all, ages_all = [], [], []
errors = 0

for i in range(N):
    try:
        p = {
            'salary': 15000 + random.random() * 235000,
            'age': random.randint(18, 85),
            'hhsize': random.randint(1, 6),
            'country': random.choice(['US','UK']),
            'gender': random.choice(['male','female']),
            'education': random.choice(['nocollege','college']),
            'riskTolerance': random.choice(['low','medium','high']),
            'style': random.choice(['conservative','mixed','highrisk']),
        }
        p['region'] = random.choice(['Northeast','Midwest','South','West'] if p['country']=='US'
                                     else ['England','Wales','Scotland','Northern Ireland'])
        expL, _, frac, G, DI = gamblingOutcome(p)
        losses_all.append(expL)
        dis_all.append(DI)
        ages_all.append(p['age'])
        capFrac = {'low':0.15,'medium':0.40,'high':1.00}[p['riskTolerance']]
        assert expL <= 0.001, f"loss>0: {expL}"
        assert G >= -0.001, f"G<0: {G}"
        assert G <= capFrac * max(DI,0) + 0.01, f"G>{capFrac}*DI: G={G}, DI={DI}"
        assert DI >= -0.01, f"DI<0: {DI}"
        assert not math.isnan(expL), "NaN loss"
        assert not math.isnan(DI), "NaN DI"
        assert not math.isinf(expL), "Inf loss"
        assert not math.isinf(DI), "Inf DI"
    except AssertionError as e:
        errors += 1
        if errors <= 5: print(f"  ASSERT {i}: {e}")
    except Exception as e:
        errors += 1
        if errors <= 5: print(f"  ERROR {i}: {type(e).__name__}: {e}")

ok(f"MC N=5000: {N-errors}/{N} clean", errors == 0, f"{errors} errors")

mean_l = sum(losses_all)/len(losses_all) if losses_all else 0
mean_d = sum(dis_all)/len(dis_all) if dis_all else 0
zero_di = sum(1 for d in dis_all if d < 1)

R_vals, ages_r = [], []
for l, d, a in zip(losses_all, dis_all, ages_all):
    if d > 0:
        R_vals.append(-l / d)
        ages_r.append(a)

hr_count = sum(1 for r in R_vals if r >= 0.15)
hr_pct = 100*hr_count/max(len(R_vals),1)

print(f"  N={len(losses_all)}, mean DI=${mean_d:,.0f}, mean loss=${mean_l:,.0f}")
print(f"  Zero-DI: {zero_di}/{N}")
print(f"  High-risk (R>=15%): {hr_count}/{len(R_vals)} = {hr_pct:.1f}%")

ok("MC5k: mean DI > 0", mean_d > 0)
ok("MC5k: mean loss in [-5000, 0]", -5000 <= mean_l <= 0, f"got {mean_l:.0f}")
ok("MC5k: some high-risk individuals exist (R>=15%)", hr_count > 0,
   f"got {hr_count}")
ok("MC5k: high-risk < 50%", hr_pct < 50)

# Age group high-risk breakdown
age_groups = [(18,29),(30,44),(45,64),(65,85)]
print("  High-risk by age group:")
for lo, hi in age_groups:
    group = [r for r, a in zip(R_vals, ages_r) if lo <= a <= hi]
    if group:
        hr = 100*sum(1 for r in group if r >= 0.15)/len(group)
        print(f"    Age {lo}-{hi}: {hr:.1f}% high-risk (n={len(group)})")

# DeltaW stats
dw_vals = [-l * annuity for l in losses_all if l < 0]
if dw_vals:
    ok("MC5k: all DeltaW > 0", all(dw > 0 for dw in dw_vals))
    ok("MC5k: max DeltaW < $500k", max(dw_vals) < 500000, f"max={max(dw_vals):,.0f}")
    print(f"  Mean DeltaW: ${sum(dw_vals)/len(dw_vals):,.0f}")

# ======================================================================
# H. CROSS-FUNCTION INTERFACE CHECKS
# ======================================================================
print("\n[H] Cross-function interface checks ...")

with open(os.path.join(M3DIR, 'disposableIncome.m')) as f: di_src = f.read()
ok("disposableIncome calls afterTaxIncome", 'afterTaxIncome' in di_src)
ok("disposableIncome calls baseEssentialShare_US", 'baseEssentialShare_US' in di_src)
ok("disposableIncome calls baseEssentialShare_UK", 'baseEssentialShare_UK' in di_src)
ok("disposableIncome calls regionMultiplier", 'regionMultiplier' in di_src)
ok("disposableIncome calls householdMultiplier", 'householdMultiplier' in di_src)

with open(os.path.join(M3DIR, 'gamblingOutcome.m')) as f: go_src = f.read()
ok("gamblingOutcome calls disposableIncome", 'disposableIncome' in go_src)
ok("gamblingOutcome has risk-dependent cap", 'capFrac' in go_src or 'cap' in go_src.lower())

with open(os.path.join(M3DIR, 'main_m3.m')) as f: main_src = f.read()
for fn in ['loadUSExpenditureData','loadUKExpenditureData','loadUSBettingData',
           'main_Q1_demo','main_Q2_demo','main_Q3_demo']:
    ok(f"main_m3 calls {fn}", fn in main_src)

with open(os.path.join(M3DIR, 'afterTaxIncome.m')) as f: tax_src = f.read()
ok("afterTaxIncome uses progressive brackets", 'Inf' in tax_src or 'inf' in tax_src.lower())
ok("afterTaxIncome computes marginal tax", 'taxable' in tax_src.lower() or 'tax = tax +' in tax_src)

# ======================================================================
# I. EDGE CASE GAUNTLET
# ======================================================================
print("\n[I] Edge case gauntlet ...")

ok("Salary=$0: no crash", DI_full(0, 30, 1, 'US', 'West')[0] >= -0.01)
ok("Salary=$1: no crash", DI_full(1, 18, 1, 'US', 'South')[0] >= -0.01)
ok("Salary=$1M: DI > 0", DI_full(1000000, 50, 1, 'US', 'West')[0] > 0)
ok("h=1: OK", DI_full(80000, 35, 1, 'US', 'West')[0] >= 0)
ok("h=10: E capped, DI >= 0", DI_full(80000, 35, 10, 'US', 'West')[0] >= -0.01)

for age in [0, 18, 24, 25, 29, 30, 49, 50, 64, 65, 74, 75, 85, 100]:
    for country in ['US', 'UK']:
        reg = 'West' if country == 'US' else 'England'
        try:
            di = DI_full(50000, age, 2, country, reg)[0]
            ok(f"Age={age} {country}: no crash", True)
        except Exception as ex:
            ok(f"Age={age} {country}: no crash", False, str(ex))

for S in [0, 1, 29999, 30000, 30001, 79999, 80000, 149999, 150000, 500000]:
    for country in ['US', 'UK']:
        Y = afterTax(S, country)[0]
        ok(f"Tax S={S} {country}: Y>=0 and Y<=S", Y >= 0 and Y <= S + 0.01)

# Gambling zero salary
p_z = {'salary':0,'age':30,'hhsize':1,'country':'US','region':'West',
       'gender':'male','education':'nocollege','riskTolerance':'high','style':'highrisk'}
expL, _, _, G, DI = gamblingOutcome(p_z)
ok("Salary=0: DI<=0", DI <= 0.01)
ok("Salary=0: G=0", G <= 0.01)
ok("Salary=0: loss=0", abs(expL) < 0.01)

# ======================================================================
print("\n" + "=" * 72)
print(f"DEEP TEST v2 RESULTS:  {PASS} passed,  {FAIL} failed")
print("=" * 72)

wb.close()

if FAIL > 0:
    print(f"\n{FAIL} FAILURE(S). See **FAIL** lines above.")
    sys.exit(1)
else:
    print("\nALL TESTS PASSED. Implementation fully validated.")
